from openpyxl import load_workbook

def process_file(file_path):
    wb=load_workbook(file_path)
    ws=wb.active

    
    total=0
    count=0

    for row in ws.iter_rows(min_row=2,values_only=True):
        value=row[1]
        count+=1
        total+=value

    average=total/count if count else 0

    return {
        "total":total,
        "average":average,
        "count":count
    }
    